import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.title("PW Validation System")

uploaded_file = st.file_uploader("Upload Excel File", type=["xlsm", "xlsx"])

if uploaded_file:

    projects_df = pd.read_excel(uploaded_file, sheet_name="Projects")
    damages_df = pd.read_excel(uploaded_file, sheet_name="Damages")
    pw_df = pd.read_excel(uploaded_file, sheet_name="PW")
    validation_df = pd.read_excel(uploaded_file, sheet_name="Validation (Vlookup)")

    validation_df = validation_df.merge(
        projects_df[['Project Number', 'Applicant', 'Disaster']],
        on='Project Number',
        how='left'
    )

    validation_df['Project Exists'] = validation_df['Applicant'].notna()

    validation_df = validation_df.merge(
        damages_df[['Damage Number', 'Project Number']],
        on='Damage Number',
        how='left',
        suffixes=('', '_Damage')
    )

    validation_df['Damage Matches Project'] = (
        validation_df['Project Number'] == validation_df['Project Number_Damage']
    )

    validation_df = validation_df.merge(
        pw_df[['PW Number', 'Disaster']],
        on='PW Number',
        how='left',
        suffixes=('', '_PW')
    )

    validation_df['PW Disaster Match'] = (
        validation_df['Disaster'] == validation_df['Disaster_PW']
    )

    validation_df['ERROR'] = ~(
        validation_df['Project Exists'] &
        validation_df['Damage Matches Project'] &
        validation_df['PW Disaster Match']
    )

    st.dataframe(validation_df)

    output_file = "Validation_Result.xlsx"
    validation_df.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active

    red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    error_col_index = list(validation_df.columns).index("ERROR") + 1

    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=error_col_index).value:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = red_fill

    wb.save(output_file)

    with open(output_file, "rb") as f:
        st.download_button("📥 Download Validated File", f, file_name=output_file)
